import os
import json
from collections import defaultdict
from openpyxl import load_workbook
from typing import Dict, Any

TARGET_ASSETS = [
    "CHEVRON",
    "CAUTIONARY_WARNING_SIGNS",
    "HAZARD",
    "PROHIBITORY_MANDATORY_SIGNS",
    "INFORMATORY_SIGNS"
]

COLUMN_MAPPING = {
    "CHEVRON": (11, 12),
    "CAUTIONARY_WARNING_SIGNS": (13, 14),
    "HAZARD": (15, 16),
    "PROHIBITORY_MANDATORY_SIGNS": (17, 18),
    "INFORMATORY_SIGNS": (19, 20),
}


# -------------------------------------------------
# JSON COUNT
# -------------------------------------------------
def _count_json(json_path: str) -> Dict[str, int]:
    counts = defaultdict(int)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    assets = data.get("assets", [])

    for asset in assets:
        asset_type = asset.get("Asset type")
        if asset_type in TARGET_ASSETS:
            counts[asset_type] += 1

    return counts


# -------------------------------------------------
# EXCEL COUNT (Furniture Chainage report)
# -------------------------------------------------
def _count_excel(excel_path: str) -> Dict[str, int]:
    counts = defaultdict(int)

    wb = load_workbook(excel_path, data_only=True)

    if "Furniture Chainage report" not in wb.sheetnames:
        raise ValueError("Furniture Chainage report sheet not found")

    ws = wb["Furniture Chainage report"]

    for row in ws.iter_rows(min_row=11):
        for asset, (col1, col2) in COLUMN_MAPPING.items():
            val1 = row[col1 - 1].value
            val2 = row[col2 - 1].value

            if isinstance(val1, (int, float)):
                counts[asset] += int(val1)

            if isinstance(val2, (int, float)):
                counts[asset] += int(val2)

    return counts


# -------------------------------------------------
# MAIN VALIDATION ENTRYPOINT
# -------------------------------------------------
def run(
    road_id: int,
    json_path: str,
    excel_path: str,
) -> Dict[str, Any]:

    json_counts = _count_json(json_path)
    excel_counts = _count_excel(excel_path)

    result = {
        "road_id": road_id,
        "status": "MATCH",
        "category_result": {},
        "total_json": sum(json_counts.values()),
        "total_excel": sum(excel_counts.values()),
    }

    for asset in TARGET_ASSETS:
        j = json_counts.get(asset, 0)
        e = excel_counts.get(asset, 0)

        match = j == e

        if not match:
            result["status"] = "MISMATCH"

        result["category_result"][asset] = {
            "json": j,
            "excel": e,
            "match": match,
        }

    if result["total_json"] != result["total_excel"]:
        result["status"] = "MISMATCH"

    return result
