# import json
# import openpyxl

# def extract_bounds(chainage_range):
#     """Extracts numeric bounds from a chainage range string (e.g., '381000 - 381500')"""
#     try:
#         start, end = map(int, chainage_range.split(" - "))
#         return min(start, end), max(start, end) 
#     except ValueError:
#         return None, None

# # Load JSON data.

# json_base = "Jsons"
# import os 
# for js in os.listdir(json_base):
#     js_path = os.path.join(json_base, js).replace("\\", "/")
#     print(js_path ,"js path ")
    
#     with open(js_path, 'r') as json_file:
#         data = json.load(json_file)


# # with open('C\:\\Users\\manav\\Desktop\\excel_folder_final_maker\\formatted_excel_script\\CR\\CR_final.json', 'r') as json_file:
    

#     # Load Excel workbook and target sheet
#     excel_file_path = 'Furniture_Chainage_Report.xlsx'
#     workbook = openpyxl.load_workbook(excel_file_path)
#     sheet = workbook['Sheet1']

#     # Define the columns where each sign type will be placed
#     columns_mapping = {
#         "Chevron": ["E", "F"],
#         "CAUTIONARY_WARNING_SIGNS": ["I", "J"],
#         "HAZARD": ["G", "H"],
#         "PROHIBITORY_MANDATORY_SIGNS": ["K", "L"],
#         "INFORMATORY_SIGNS": ["M", "N", "O"]
#     }

#     for road_section, chainages in data.items():
#         for chainage_range, details in chainages.items():
#             from_chainage = details['from']
#             to_chainage = details['to']
#             match_string = f"{from_chainage} - {to_chainage}"
#             reverse_match_string = f"{to_chainage} - {from_chainage}"

#             prev_chainage = None  
#             for row in range(7, sheet.max_row + 1):  
#                 excel_chainage = sheet[f'A{row}'].value  
#                 excel_road_section = sheet[f'B{row}'].value  

#                 if excel_chainage is None:
#                     excel_chainage = prev_chainage
#                 else:
#                     prev_chainage = excel_chainage  

#                 if excel_chainage is None:
#                     continue

#                 excel_start, excel_end = extract_bounds(excel_chainage)
#                 match_start, match_end = extract_bounds(match_string)
                
#                 # Check if both chainage range and road section match (including reverse chainage order)
#                 if (excel_road_section == road_section and
#                     ((excel_chainage == match_string or excel_chainage == reverse_match_string) or
#                     (excel_start <= match_start <= excel_end and excel_start <= match_end <= excel_end))):
                    
#                     # Populate data into the matched row
#                     sheet[f'C{row}'].value = from_chainage
#                     sheet[f'D{row}'].value = to_chainage

                    
#                     for asset_type, columns in columns_mapping.items():
#                         if asset_type in details:
#                             asset_values = details[asset_type]
                            
#                             # Populate Avenue/Left and Median/Right for each asset
#                             if len(columns) >= 2:
#                                 sheet[f'{columns[0]}{row}'].value = asset_values.get('Avenue/Left', 0)
#                                 sheet[f'{columns[1]}{row}'].value = asset_values.get('Median/Right', 0)
                            
#                             # Populate Overhead Signs if present
#                             if len(columns) == 3 and "Overhead Signs" in asset_values:
#                                 sheet[f'{columns[2]}{row}'].value = asset_values["Overhead Signs"]

#                     break

#     # Save the updated workbook to a new file
#     updated_excel_file_path = 'Bambaore_garamore_final.xlsx'
#     workbook.save(updated_excel_file_path)

#     print(f"Updated data saved to {updated_excel_file_path}")
import json
import openpyxl
import os
from django.conf import settings



def extract_bounds(chainage_range):
    """Extracts numeric bounds from a chainage range string (e.g., '381000 - 381500')"""
    try:
        start, end = map(int, chainage_range.split(" - "))
        return min(start, end), max(start, end)
    except ValueError:
        return None, None


def excel_updation(excel_name,session_folder_path):
    base_path = session_folder_path
    # Load Excel workbook once
    excel_file_path = 'Furniture_Chainage_Report.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook['Sheet1']

    # Define the columns where each sign type will be placed
    columns_mapping = {
        "Chevron": ["E", "F"],
        "CAUTIONARY_WARNING_SIGNS": ["I", "J"],
        "HAZARD": ["G", "H"],
        "PROHIBITORY_MANDATORY_SIGNS": ["K", "L"],
        "INFORMATORY_SIGNS": ["M", "N", "O"]
    }

    # Process all JSON files in the folder
    json_base = os.path.join(base_path , "Jsons" )
    for js in os.listdir(json_base):
        js_path = os.path.join(json_base, js).replace("\\", "/")
        print(f"Processing JSON file: {js_path}")

        with open(js_path, 'r') as json_file:
            data = json.load(json_file)

        for road_section, chainages in data.items():
            for chainage_range, details in chainages.items():
                from_chainage = details['from']
                to_chainage = details['to']
                match_string = f"{from_chainage} - {to_chainage}"
                reverse_match_string = f"{to_chainage} - {from_chainage}"

                prev_chainage = None
                for row in range(7, sheet.max_row + 1):
                    excel_chainage = sheet[f'A{row}'].value
                    excel_road_section = sheet[f'B{row}'].value

                    if excel_chainage is None:
                        excel_chainage = prev_chainage
                    else:
                        prev_chainage = excel_chainage

                    if excel_chainage is None:
                        continue

                    excel_start, excel_end = extract_bounds(excel_chainage)
                    match_start, match_end = extract_bounds(match_string)

                    # Check if both chainage range and road section match (including reverse chainage order)
                    if (excel_road_section == road_section and
                        ((excel_chainage == match_string or excel_chainage == reverse_match_string) or
                        (excel_start <= match_start <= excel_end and excel_start <= match_end <= excel_end))):

                        # Populate data into the matched row
                        sheet[f'C{row}'].value = from_chainage
                        sheet[f'D{row}'].value = to_chainage

                        for asset_type, columns in columns_mapping.items():
                            if asset_type in details:
                                asset_values = details[asset_type]

                                # Populate Avenue/Left and Median/Right for each asset
                                if len(columns) >= 2:
                                    sheet[f'{columns[0]}{row}'].value = asset_values.get('Avenue/Left', 0)
                                    sheet[f'{columns[1]}{row}'].value = asset_values.get('Median/Right', 0)

                                # Populate Overhead Signs if present
                                if len(columns) == 3 and "Overhead Signs" in asset_values:
                                    sheet[f'{columns[2]}{row}'].value = asset_values["Overhead Signs"]

    totals = {col: 0 for cols in columns_mapping.values() for col in cols}
    for col in totals.keys():
        for row in range(7, sheet.max_row + 1):
            cell_value = sheet[f'{col}{row}'].value
            if isinstance(cell_value, (int, float)):
                totals[col] += cell_value
            elif isinstance(cell_value, str) and cell_value.isdigit():
                totals[col] += int(cell_value)


    total_row = sheet.max_row + 2
    sheet[f'A{total_row}'].value = "Total Counts"
    for col, total in totals.items():
        sheet[f'{col}{total_row}'].value = total

    final_total = sum(totals.values())
    sheet[f'A{total_row + 1}'].value = "Grand Total"
    sheet[f'B{total_row + 1}'].value = final_total

    # Save the updated workbook to a new file
    updated_excel_file_path =  os.path.join(base_path, f'generated_files/{excel_name}.xlsx')
    workbook.save(updated_excel_file_path)

    print(f"Updated data saved to {updated_excel_file_path}")

