import os
import openpyxl

TARGET_ASSETS = [
    "CHEVRON",
    "CAUTIONARY_WARNING_SIGNS",
    "HAZARD",
    "PROHIBITORY_MANDATORY_SIGNS",
    "INFORMATORY_SIGNS"
]


# ==========================================
# FIND FINAL EXCEL
# ==========================================

def find_final_excel(final_output_dir, logger):

    if not os.path.exists(final_output_dir):
        logger.error("Final_Output folder not found")
        return None

    for file in os.listdir(final_output_dir):

        if file.startswith("Final Excel") and file.endswith(".xlsx"):
            file_path = os.path.join(final_output_dir, file)

            logger.info(f"Final Excel found: {file_path}")

            return file_path

    logger.error("Final Excel file not found")

    return None


# ==========================================
# READ COUNTS FROM EXCEL
# ==========================================

def read_counts(file_path, logger):

    try:

        wb = openpyxl.load_workbook(file_path, data_only=True)

        sheet = wb.active

        total_row = None

        # find "Total" row
        for row in range(1, sheet.max_row + 1):

            value = sheet.cell(row=row, column=1).value

            if value and "Total" in str(value):

                total_row = row

                break

        if not total_row:

            logger.error("Total row not found in Excel")

            return None

        # read counts
        chevron = (sheet.cell(total_row, 5).value or 0) + \
            (sheet.cell(total_row, 6).value or 0)

        hazard = (sheet.cell(total_row, 7).value or 0) + \
            (sheet.cell(total_row, 8).value or 0)

        caution = (sheet.cell(total_row, 9).value or 0) + \
            (sheet.cell(total_row, 10).value or 0)

        prohibitory = (sheet.cell(total_row, 11).value or 0) + \
            (sheet.cell(total_row, 12).value or 0)

        informatory = (sheet.cell(total_row, 13).value or 0) + \
            (sheet.cell(total_row, 14).value or 0)

        counts = {

            "CHEVRON": chevron,

            "CAUTIONARY_WARNING_SIGNS": caution,

            "HAZARD": hazard,

            "PROHIBITORY_MANDATORY_SIGNS": prohibitory,

            "INFORMATORY_SIGNS": informatory

        }

        return counts

    except Exception as e:

        logger.exception(f"Failed reading Excel: {e}")

        return None


# ==========================================
# MODULE ENTRYPOINT
# ==========================================

def run(survey_root, logger):

    logger.info("XLSX VALIDATOR STARTED")

    final_output_dir = os.path.join(survey_root, "Final_Output")

    file_path = find_final_excel(final_output_dir, logger)

    if not file_path:

        return None

    counts = read_counts(file_path, logger)

    if not counts:

        return None

    total = sum(counts.values())

    logger.info("========== XLSX COUNTS ==========")

    for k, v in counts.items():

        logger.info(f"{k} = {v}")

    logger.info(f"TOTAL ASSETS = {total}")

    logger.info("=================================")

    return {
        "categories": counts,
        "total_assets": total
    }
