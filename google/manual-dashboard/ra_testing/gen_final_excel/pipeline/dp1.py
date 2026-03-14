import re
import json
import openpyxl
import os
import requests
import logging

logger = logging.getLogger(__name__)


def extract_bounds(chainage_range):
    try:
        start, end = map(int, chainage_range.split(" - "))
        return min(start, end), max(start, end)
    except ValueError:
        return None, None


def normalize_string(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def ranges_overlap(a_start, a_end, b_start, b_end):
    return not (a_end < b_start or a_start > b_end)


# def get_project_name_from_roads(survey_id, logger):

#     try:
#         url = f"https://ndd.roadathena.com/api/surveys/{survey_id}"

#         response = requests.get(
#             url,
#             headers={"security-Password": "admin@123"},
#             timeout=10
#         )

#         if response.status_code != 200:
#             logger.warning("Failed to fetch survey roads")
#             return f"Survey_{survey_id}"

#         data = response.json()
#         roads = data.get("roads", [])

#         if not roads:
#             return f"Survey_{survey_id}"

#         pattern = r"(MCW\s*LHS|MCW\s*RHS|SRR\s*\d+|SRL\s*\d+|TR\s*\d+|TL\s*\d+|IRR\s*\d+|IRL\s*\d+|CR|CL|LRR|LRL)\s*$"

#         clean_names = []

#         for road in roads:
#             name = road.get("road_name", "")
#             cleaned = re.sub(pattern, "", name, flags=re.IGNORECASE).strip()
#             cleaned = cleaned.replace("/", "")
#             clean_names.append(cleaned)

#         project_name = max(
#             clean_names, key=len) if clean_names else f"Survey_{survey_id}"

#         return project_name

#     except Exception as e:
#         logger.warning(f"Project name extraction failed: {e}")
#         return f"Survey_{survey_id}"


# Load Excel workbook once
# excel_file_path = 'Furniture_Chainage_Report.xlsx'
# workbook = openpyxl.load_workbook(excel_file_path)
# sheet = workbook['Sheet1']

columns_mapping = {
    "CHEVRON": ["E", "F"],
    "CAUTIONARY_WARNING_SIGNS": ["I", "J"],
    "HAZARD": ["G", "H"],
    "PROHIBITORY_MANDATORY_SIGNS": ["K", "L"],
    "INFORMATORY_SIGNS": ["M", "N", "O"]
}


def run(survey_root, survey_id, logger):

    logger.info("Starting dp1 asset mapping")
    logger.debug(
        f"dp1.run called with survey_root={survey_root}, survey_id={survey_id}")

    # -----------------------------
    # PATHS
    # -----------------------------

    json_base = os.path.join(survey_root, "jsons")
    logger.debug(f"Looking for JSONs in {json_base}")
    if not os.path.exists(json_base):
        logger.error(f"Json folder not found: {json_base}")
        return None

    # primary location: same directory as this module
    excel_file_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),  # goes up from pipeline/ to gen_final_excel/
        "Furniture_Chainage_Report.xlsx"
    )

    logger.debug(f"Template excel path: {excel_file_path}")

    # fallback to workspace root (where master_runner resides)
    if not os.path.exists(excel_file_path):
        alt = os.path.join(os.getcwd(), "Furniture_Chainage_Report.xlsx")
        logger.debug(f"Primary template missing, trying workspace root: {alt}")
        if os.path.exists(alt):
            excel_file_path = alt
        else:
            logger.error(f"Template Excel not found: {excel_file_path}")
            return None

    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook['Sheet1']

    logger.info(f"Loaded template Excel: {excel_file_path}")
    logger.debug(f"Workbook sheets: {workbook.sheetnames}")


# json_base = "jsons"

    for js in os.listdir(json_base):

        if not js.endswith(".json"):
            continue

        js_path = os.path.join(json_base, js).replace("\\", "/")
        logger.info(f"Processing JSON file: {js_path}")
        logger.debug(f"Reading data from {js_path}")

        with open(js_path, 'r') as json_file:
            data = json.load(json_file)

        for road_section, chainages in data.items():
            logger.debug(
                f"Road section: {road_section} contains {len(chainages)} ranges")
            for chainage_range, details in chainages.items():

                from_chainage = details['from']
                to_chainage = details['to']
                logger.debug(
                    f"Chainage range {chainage_range} -> from {from_chainage} to {to_chainage}")

                json_start, json_end = min(from_chainage, to_chainage), max(
                    from_chainage, to_chainage)

                prev_chainage = None

                for row in range(7, sheet.max_row + 1):

                    excel_chainage = sheet[f'A{row}'].value
                    excel_road_section = sheet[f'B{row}'].value
                    if logger:
                        logger.debug(
                            f"Row {row} excel_chainage={excel_chainage} excel_road_section={excel_road_section}")

                    if excel_chainage is None:
                        excel_chainage = prev_chainage
                    else:
                        prev_chainage = excel_chainage

                    if excel_chainage is None:
                        continue

                    excel_start, excel_end = extract_bounds(excel_chainage)
                    if excel_start is None:
                        continue

                    overlap = ranges_overlap(
                        json_start, json_end, excel_start, excel_end)

                    # ✅ NORMALIZED ROAD SECTION MATCH
                    if (
                        normalize_string(excel_road_section) ==
                        normalize_string(road_section)
                        and ranges_overlap(json_start, json_end, excel_start, excel_end)
                    ):
                        logger.debug(
                            f"Match found on row {row} for {road_section} {chainage_range}")

                        # Populate data into matched row
                        sheet[f'C{row}'].value = from_chainage
                        sheet[f'D{row}'].value = to_chainage

                        for asset_type, columns in columns_mapping.items():

                            if asset_type in details:
                                asset_values = details[asset_type]

                                # Avenue/Left (ACCUMULATE)
                                if len(columns) >= 1:
                                    current_val = sheet[f'{columns[0]}{row}'].value or 0
                                    sheet[f'{columns[0]}{row}'].value = current_val + \
                                        asset_values.get('Avenue/Left', 0)

                                # Median/Right (ACCUMULATE)
                                if len(columns) >= 2:
                                    current_val = sheet[f'{columns[1]}{row}'].value or 0
                                    sheet[f'{columns[1]}{row}'].value = current_val + \
                                        asset_values.get('Median/Right', 0)

                                # Overhead Signs (ACCUMULATE safely)
                                # if len(columns) == 3:
                                #     overhead_val = asset_values.get(
                                #         "Overhead Signs", 0)

                                #     if isinstance(overhead_val, str):
                                #         if overhead_val.strip().upper() == "NONE":
                                #             overhead_val = 0
                                #         elif overhead_val.isdigit():
                                #             overhead_val = int(overhead_val)
                                #         else:
                                #             overhead_val = 0

                                #     current_val = sheet[f'{columns[2]}{row}'].value or 0
                                #     sheet[f'{columns[2]}{row}'].value = current_val + \
                                #         overhead_val
                                if len(columns) == 3 and "Overhead Signs" in asset_values:
                                    sheet[f'{columns[2]}{row}'].value = asset_values["Overhead Signs"]

                        break   # 🔥 VERY IMPORTANT — avoid duplicate match


# =========================
# TOTAL CALCULATION (UNCHANGED STRUCTURE)
# =========================

    totals = {col: 0 for cols in columns_mapping.values() for col in cols}
    logger.debug(
        f"Starting total calculation for columns: {list(totals.keys())}")

    for col in totals.keys():
        for row in range(7, sheet.max_row + 1):
            cell_value = sheet[f'{col}{row}'].value
            if isinstance(cell_value, (int, float)):
                totals[col] += cell_value
            elif isinstance(cell_value, str) and cell_value.isdigit():
                totals[col] += int(cell_value)

    total_row = sheet.max_row + 2
    sheet[f'A{total_row}'].value = "Total Counts"

    for col, total in totals.items():
        sheet[f'{col}{total_row}'].value = total

    final_total = sum(totals.values())
    logger.debug(f"Calculated final_total={final_total}")

    sheet[f'A{total_row + 1}'].value = "Grand Total"
    sheet[f'B{total_row + 1}'].value = final_total

    # -----------------------------
# GET PROJECT NAME
# -----------------------------

    # project_name = get_project_name_from_roads(survey_id, logger)

    # project_name = project_name.replace("/", "").strip()

    # -----------------------------
    # OUTPUT FOLDER (use master_runner Final_Output)
    # -----------------------------

    # final_folder = os.path.join(
    #     survey_root,
    #     "Final_Output"
    # )

    # os.makedirs(final_folder, exist_ok=True)

    # # Use master_runner's standardized final filename
    # safe_project_name = project_name.replace("/", "").strip()

    # updated_excel_file_path = os.path.join(
    #     final_folder,
    #     f"{safe_project_name}_Final_Report.xlsx"
    # )

    # workbook.save(updated_excel_file_path)

    # logger.info(f"Final Excel saved: {updated_excel_file_path}")
    # logger.info(f"Grand Total: {final_total}")
    # return updated_excel_file_path
    return workbook
