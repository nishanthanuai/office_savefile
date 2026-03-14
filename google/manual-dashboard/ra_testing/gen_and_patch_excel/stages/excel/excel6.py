from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl import Workbook, load_workbook
from collections import defaultdict
import math,bisect,json,os,requests,logging
from core.sign_maps import SIGN_CATEGORY_MAP
from typing import Dict, Any, Optional
logger = logging.getLogger(__name__)


# 🔹 ADD THIS HERE (global constant)
CATEGORY_COLUMNS = [
    ("CHEVRON", "K", "L"),
    ("CAUTIONARY_WARNING_SIGNS", "M", "N"),
    ("HAZARD", "O", "P"),
    ("PROHIBITORY_MANDATORY_SIGNS", "Q", "R"),
    ("INFORMATORY_SIGNS", "S", "T"),
]
# Pre-normalized map — built ONCE at import time for O(1) lookup
NORMALIZED_SIGN_MAP = {
    k.strip().upper().replace(" ", "_"): v
    for k, v in SIGN_CATEGORY_MAP.items()
}

def get_max_distance(gpx_data):
    max_value = max((x['distanceInMeters']
                    for x in gpx_data.values()), default=0)
    # max_value = ((max_value + 499) // 500) * 500
    return math.ceil(max_value)


def get_road_type(road_name: str):
    road_name = road_name.upper()

    SERVICE_KEYWORDS = ["SRR", "SRL", "IRR", "IRL", "TR", "TL", "LRL", "LRR"]

    if any(key in road_name for key in SERVICE_KEYWORDS):
        return "SERVICE"
    elif "MCW" in road_name:
        return "MCW"
    else:
        return "UNKNOWN"


def get_informatory_count(section_data):
    return (
        section_data.get("INFORMATORY_SIGNS", 0) +
        section_data.get("DIGITAL_SPEED_DISPLAY_SIGNS", 0) +
        section_data.get("VARIABLE_MESSAGE_SIGNS", 0)
    )


# get closest timestamp using fast binary search O(log n)
def find_closest_timestamp(target_distance, sorted_distances, sorted_gpx, max_distance_difference=500):
    idx = bisect.bisect_left(sorted_distances, target_distance)
    best = None
    best_diff = float('inf')
    # Check neighbors in sorted order (idx-1, idx, idx+1)
    for i in range(max(0, idx - 1), min(len(sorted_gpx), idx + 2)):
        diff = abs(sorted_distances[i] - target_distance)
        if diff <= max_distance_difference and diff < best_diff:
            best_diff = diff
            best = sorted_gpx[i]
    return best

def generate_counts_dict():
    """Initializes a dictionary with 0 for all required Excel categories."""
    return {
        "CHEVRON": 0,
        "CAUTIONARY_WARNING_SIGNS": 0,
        "HAZARD": 0,
        "PROHIBITORY_MANDATORY_SIGNS": 0,
        "INFORMATORY_SIGNS": 0,
        "DIGITAL_SPEED_DISPLAY_SIGNS": 0,
        "VARIABLE_MESSAGE_SIGNS": 0
    }

def process_item(counts, item, side_key):
    """Maps specific signs to categories and updates counts."""
    if side_key not in counts:
        counts[side_key] = generate_counts_dict()

    raw_key = item.get("Asset type") or item.get("Anomaly type")
    if not raw_key:
        return counts
    normalized_key = raw_key.strip().upper().replace(" ", "_")

    category = NORMALIZED_SIGN_MAP.get(normalized_key)

    if category:
        counts[side_key][category] += 1
    else:
        if raw_key not in counts[side_key]:
            counts[side_key][raw_key] = 0
        counts[side_key][raw_key] += 1

    return counts

def add_gps_data(result, key, sorted_distances, sorted_gpx, distance_start, distance_end):
    """Add GPS data to the result for the given range using fast lookup."""
    closest_from_values = find_closest_timestamp(distance_start, sorted_distances, sorted_gpx)
    closest_to_values = find_closest_timestamp(distance_end, sorted_distances, sorted_gpx)

    if closest_from_values is not None:
        result[key]['lat'] = closest_from_values.get('lat', 0)
        result[key]['lng'] = closest_from_values.get('lng', 0)

    if closest_to_values is not None:
        result[key]['to_lat'] = closest_to_values.get('lat', 0)
        result[key]['to_lng'] = closest_to_values.get('lng', 0)



def parse_json(json_data, max_distance, gpx_data, road_name, first_gap, interval=500):
    result = {}

    # 🔹 1. Pre-sort GPX Data ONCE for fast binary search
    sorted_gpx = sorted(gpx_data.values(), key=lambda v: float(v.get('distanceInMeters', 0)))
    sorted_distances = [v['distanceInMeters'] for v in sorted_gpx]

    assets = json_data.get("assets") or []
    anomalies = json_data.get("anomalies") or []

    if not isinstance(assets, list):
        assets = []
    if not isinstance(anomalies, list):
        anomalies = []

    all_items = assets + anomalies
    road_type = get_road_type(road_name)

    # 🔹 2. Build the distance ranges and initialize the result dictionary
    distance_ranges = []
    start = 0

    if max_distance <= 0:
        max_distance = max(
            [float(x.get("Distance", 0) or 0) for x in all_items],
            default=0
        )

    while start < max_distance:
        end = first_gap if start == 0 else min(start + interval, max_distance)
        distance_ranges.append((start, end))
        
        # Initialize the dictionaries for this range
        if road_type == "SERVICE":
            result[start] = {"Left": generate_counts_dict(), "Right": generate_counts_dict()}
        else:
             result[start] = {"Avenue": generate_counts_dict(), "Median": generate_counts_dict(), "Center": generate_counts_dict()}
        
        # Pre-fill GPS data using the fast lookup
        add_gps_data(result, start, sorted_distances, sorted_gpx, start, end)
        
        start = end

    # 🔹 3. ONE PASS mathematically bucket over all items
    for item in all_items:
        try:
            distance = float(item.get("Distance") or 0)
        except Exception as e:
            continue
        
        raw_side = (item.get("Side") or "").strip().title()

        # Find which bucket this item belongs to mathematically
        if distance < first_gap:
            bucket_key = 0
        else:
            bucket_key = first_gap + ((int(distance) - first_gap) // interval) * interval

        if bucket_key not in result:
            continue
        
        # --- SIDE NORMALIZATION ---
        if road_type == "SERVICE":
            side = raw_side if raw_side in ["Left", "Right"] else "Left"
        else:  # MCW
            if raw_side in ["Avenue", "Median"]:
                side = raw_side
            elif raw_side == "Left":
                side = "Avenue"
            elif raw_side == "Right":
                side = "Median"
            else:
                continue  # skip unknown MWC sides
                
        # --- CATEGORY NORMALIZATION ---
        raw_key = (
            item.get("Asset type")
            or item.get("Anomaly type")
            or ""
        ).strip()

        if not raw_key:
            continue

        normalized_key = raw_key.strip().upper().replace(" ", "_")
        category = NORMALIZED_SIGN_MAP.get(normalized_key)

        if not category:
            logger.warning("Unmapped category skipped: %s", raw_key)
            continue

        # Add to correct bucket safely
        result[bucket_key][side][category] += 1

    return result



def previous_value_divisible_by_500(value):
    if value < 500:
        return 0
    return (value // 500) * 500


def previous_next_divisible_by_500(value):

    previous_value = (value // 500) * 500
    next_value = previous_value + 500
    return next_value


# Function to handle sorting of the assets based on distance
def sort_assets_by_distance(data):
    # Sort the assets, placing None (null) distances at the end of the list
    sorted_assets = sorted(
        data['assets'],
        key=lambda asset: (asset['Distance'] is None, asset['Distance'])
    )
    return sorted_assets


def run(output_json_path: str, gpx_json_path: str, output_folder: str, road_id: int, road_data: Dict[str, Any], logger: Optional[logging.Logger] = None,) -> Dict[str, Any]:
    logger = logger or logging.getLogger(__name__)

    # def apply_common_chainage_formatting(sheet, survey_data):
    #     thick_border = Border(left=Side(style='medium'),
    #                           right=Side(style='medium'),
    #                           top=Side(style='medium'),
    #                           bottom=Side(style='medium'))
    #     # print(road_data)
    #     print()

    #     # A1 - Title
    #     # sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nMC: {survey_data['mc']['name']}\nSubdivision: {survey_data['sub_division']['sub_division']}"
    #     sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {road_data['road']['ho']['name']}\nPIU : {road_data['road']['ro']['name']}"
    #     # sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {'Gandinagar'}\nPIU : {'Rajpur'}"
    #     sheet['A1'].font = Font(b=True, size=14)
    #     sheet['A1'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    #     sheet.merge_cells('A1:L1')
    #     sheet['A1'].border = thick_border

    def apply_common_chainage_formatting(sheet, survey_data):
        thick_border = Border(left=Side(style='medium'),
                              right=Side(style='medium'),
                              top=Side(style='medium'),
                              bottom=Side(style='medium'))

        # A1 - Title
        sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {road_data['road']['ho']['name']}\nPIU : {road_data['road']['ro']['name']}"
        sheet['A1'].font = Font(b=True, size=14)
        sheet['A1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        sheet.merge_cells('A1:L1')  # Merge cells A1 to L1

        # Apply border to the merged range A1:L1
        for col in range(1, 13):  # Columns A to L
            cell = sheet.cell(row=1, column=col)
            cell.border = thick_border

        # A2 - Date of survey
        sheet["A2"] = f"Date of survey: {survey_data['survey_date']}"
        sheet['A2'].font = Font(b=True, size=13)
        sheet.merge_cells('A2:G2')
        sheet['A2'].alignment = Alignment(
            horizontal='center', vertical='center')
        sheet['A2'].border = thick_border

        for col in range(1, 8):  # Columns A to L
            cell = sheet.cell(row=2, column=col)
            cell.border = thick_border

        # A3 - Segment name
        sheet["A3"] = f"Segment name: {survey_data['road_name']}"
        sheet.merge_cells('A3:G3')
        sheet['A3'].font = Font(size=13, b=True)
        sheet['A3'].alignment = Alignment(
            horizontal='center', vertical='center')
        sheet['A3'].border = thick_border

        for col in range(1, 8):  # Columns A to L
            cell = sheet.cell(row=3, column=col)
            cell.border = thick_border

        # A4 - Start Chainage
        sheet["A4"] = f"Start Chainage"
        sheet.merge_cells('A4:B4')
        sheet['A4'].font = Font(size=13, b=True)
        sheet['A4'].alignment = Alignment(
            horizontal='center', vertical='center')
        sheet['A4'].border = thick_border

        for col in range(1, 4):  # Columns A to L
            cell = sheet.cell(row=4, column=col)
            cell.border = thick_border

        # B4 - Start Chainage value
        sheet["C4"] = f"{survey_data['start_chainage']}"
        sheet['C4'].font = Font(size=13)
        sheet.merge_cells('C4:D4')
        sheet['C4'].alignment = Alignment(
            horizontal='center', vertical='center')
        sheet['C4'].border = thick_border
        for col in range(1, 4):
            cell = sheet.cell(row=4, column=3)
            cell.border = thick_border

        # A5 - End Chainage
        sheet["A5"] = f"End Chainage"
        sheet.merge_cells('A5:B5')
        sheet['A5'].font = Font(size=13, b=True)
        sheet['A5'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        sheet['A5'].border = thick_border

        for col in range(1, 4):
            cell = sheet.cell(row=5, column=col)
            cell.border = thick_border

        # B5 - End Chainage value
        # sheet["C5"] = f"{survey_data['end_chainage']}"
        # sheet['C5'].font = Font(size=13)
        # sheet.merge_cells('C5:D5')
        # sheet['C5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        # sheet['C5'].border = thick_border

        end_chainage_value = survey_data.get('end_chainage', 'N/A')
        sheet["C5"] = f"{end_chainage_value}"
        sheet['C5'].font = Font(size=13)
        sheet.merge_cells('C5:D5')
        sheet['C5'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        sheet['C5'].border = thick_border
        for col in range(1, 4):
            cell = sheet.cell(row=5, column=3)
            cell.border = thick_border

        logger.debug("End chainage value retrieved: %s", end_chainage_value)

        # Adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.row_dimensions[1].height = 80
        sheet.row_dimensions[2].height = 30
        sheet.row_dimensions[3].height = 30
        sheet.row_dimensions[4].height = 50
        sheet.row_dimensions[5].height = 50
        sheet.row_dimensions[6].height = 50

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20

        return sheet

    def countingExcelReport(survey_data):

        # wb = Workbook()
        # wb = load_workbook(os.path.join(output_folder, f"{road_data['road']['name']}_formatted.xlsx"))
        wb = load_workbook(os.path.join(
            output_folder, f"{road_id}_formatted.xlsx"))

        # report_name = 'Road Furniture Chainage wise report'
        report_name = 'Furniture Chainage report'
        if report_name in wb.sheetnames:

            del wb[report_name]

        ws = wb.create_sheet(report_name)
        ws.title = report_name
        wb.active = ws
        ws = apply_common_chainage_formatting(ws, survey_data)
        # print(ws)

        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        road_type = get_road_type(road_name)
        if road_type == "SERVICE":
            headers = {
                "A7": "Chainage", "A8": "From", "B8": "To",
                "C8": "From", "E8": "To", "C7": "Geo-Location(Lat,Long)",
                "G7": "Survey Name", "H7": "Survey Date", "I7": "Direction",
                "J7": "Side", "K7": "Furniture Assets", "K8": "CHEVRON", "K10": "Left", "L10": "Right",
                "M8": "CAUTIONARY_WARNING_SIGNS", "M10": "Left", "N10": "Right",
                "O8": "HAZARD", "O10": "Left", "P10": "Right",
                "Q8": "PROHIBITORY_MANDATORY_SIGNS", "Q10": "Left", "R10": "Right",
                "S8": "INFORMATORY_SIGNS", "S10": "Left", "T10": "Right", "A6": report_name,
            }
        else:
            headers = {
                "A7": "Chainage", "A8": "From", "B8": "To",
                "C8": "From", "E8": "To", "C7": "Geo-Location(Lat,Long)",
                "G7": "Survey Name", "H7": "Survey Date", "I7": "Direction",
                "J7": "Side", "K7": "Furniture Assets", "K8": "CHEVRON", "K10": "Avenue", "L10": "Median",
                "M8": "CAUTIONARY_WARNING_SIGNS", "M10": "Avenue", "N10": "Median", "O8": "HAZARD", "O10": "Avenue", "P10": "Median", "Q8": "PROHIBITORY_MANDATORY_SIGNS",
                "Q10": "Avenue", "R10": "Median", "S8": "INFORMATORY_SIGNS", "S10": "Avenue", "T10": "Median", "A6": report_name,
            }

        for cell, value in headers.items():
            ws[cell] = value

        merge_ranges = [
            ("A6", "O6"), ("A7", "B7"), ("C7", "F7"),
            ("G7", "G9"), ("H7", "H9"), ("I7", "I9"), ("J7", "J9"),
            ("K7", "T7"), ("A8", "A9"), ("B8", "B9"), ("C8", "D9"),
            ("E8", "F9"), ("K8", "L9"), ("M8", "N9"), ("O8",
                                                       "P9"), ('Q8', 'R9'), ('S8', 'T9'), ('U8', 'V9')
        ]

        COLOR = "6AC9FF"
        for start, end in merge_ranges:
            ws.merge_cells(f"{start}:{end}")

        for row in range(6, 11):
            for col in "ABCDEFGHIJKLMNOPQRST":
                cell = ws[f'{col}{row}']
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.font = Font(name="Arial", bold=True)
                cell.fill = PatternFill(
                    start_color=COLOR, end_color=COLOR, fill_type="solid")
                cell.border = thin_border

        column_widths = {
            'A': 15, 'B': 15, 'C': 16, 'D': 16, 'E': 16,
            'F': 16, 'G': 25, 'H': 15, 'I': 25, 'J': 12,
            'K': 12, 'L': 12, 'M': 16, 'N': 12, 'O': 16, 'P': 16, 'Q': 16, 'R': 16, 'S': 16, 'T': 16
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        row_heights = {1: 55, 2: 40, 3: 40, 4: 40}

        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        return wb

    def chainageWiseCounting(survey_data, json_data, side, value_diff):

        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        logger.debug("Survey side fetched: %s", survey_data.get("side"))
        # side = survey_data.get("side")
        # side = "RHS"
        # 🔹 Decide sides dynamically (ONCE per road)
        road_type = get_road_type(road_name)

        if road_type == "SERVICE":
            sides = ("Left", "Right")
        else:
            sides = ("Avenue", "Median")

        if side == "LHS":
            survey_direction = "Increasing"
        elif side == "RHS":
            survey_direction = "Decreasing"
        else:
            survey_direction = "None"
        logger.debug("Side value (temporary debug): %s", side)

        if survey_data["start_chainage"]:

            # start_chainage = int(survey_data["start_chainage"].split("+")[0]) * 1000 + int(survey_data["start_chainage"].split("+")[1])
            start_chainage = int(survey_data["start_chainage"])

        else:
            start_chainage = 0

        survey_date = survey_data["survey_date"]
        survey_name = survey_data["survey_name"]
        survey_direction = survey_direction
        survey_side = side
        logger.debug("Survey side: %s", survey_side)

        wb = countingExcelReport(survey_data)
        report_name = 'Furniture Chainage report'
        ws = wb[report_name]
        wb.active = ws

        start_row = 11
        to_chainage = 0

        end_chainage_value = survey_data.get('end_chainage', None)

        chainage_diff = value_diff

        # if chainage_diff > 0:
        #     side = "RHS"
        # else :
        #     side = "LHS"

        # start_gap = start_chainage

        if end_chainage_value is None:
            logger.warning(
                "End chainage value is not defined. Exiting process.")
        else:
            try:
                end_chainage_value = int(end_chainage_value)
                logger.debug(
                    "Validated end chainage value: %d",
                    end_chainage_value)
            except ValueError:
                logger.error(
                    "Invalid end chainage value: %s. Must be a valid integer.",
                    end_chainage_value)

        logger.debug(
            "Chainage side value: %s",
            side
        )
        for i, (key, data) in enumerate(json_data.items()):

            current_row = start_row + i

            if i == 0:

                if side == "RHS":

                    from_chainage = int(start_chainage)

                    if start_chainage != 0 and int(start_chainage) % 500 == 0:
                        to_chainage = from_chainage - 500

                    if start_chainage != 0 and int(start_chainage) % 500 != 0:
                        logger.debug("Entered condition: true branch")

                        to_chainage = previous_value_divisible_by_500(
                            int(start_chainage))

                if side == "LHS":

                    from_chainage = int(start_chainage)

                    if start_chainage != 0 and int(start_chainage) % 500 == 0:
                        to_chainage = from_chainage + 500

                    if start_chainage != 0 and int(start_chainage) % 500 != 0:

                        to_chainage = previous_next_divisible_by_500(
                            int(start_chainage))
            else:

                if from_chainage == start_chainage and side == "RHS":

                    from_chainage = to_chainage
                    to_chainage = from_chainage - 500

                elif side == "RHS":

                    from_chainage = from_chainage - 500
                    to_chainage = from_chainage - 500
                else:
                    from_chainage = to_chainage
                    to_chainage = to_chainage + 500
                # Check if to_chainage exceeds the chainage_limit
                if side == "LHS":

                    if end_chainage_value:
                        if to_chainage > end_chainage_value and to_chainage - end_chainage_value < 500:
                            to_chainage = end_chainage_value

                        if to_chainage > end_chainage_value:
                            logger.warning(
                                "Chainage limit exceeded: %s. Stopping further processing.",
                                end_chainage_value)
                            break
                else:
                    if side == "RHS":
                        if to_chainage < 0:
                            logger.warning(
                                "Chainage limit exceeded: %s. Stopping further processing.",
                                end_chainage_value
                            )
                            break

                        if i == len(json_data) - 1:
                            to_chainage = end_chainage_value
                    else:

                        if to_chainage == end_chainage_value:
                            logger.warning(
                                "Chainage limit exceeded: %s. Stopping further processing.",
                                end_chainage_value
                            )
                            break

            logger.debug(
                "Chainage range | start=%s | from=%s | to=%s",
                start_chainage,
                from_chainage,
                to_chainage
            )

            try:
                if i == 0:
                    data2 = json_data[(chainage_diff)]
                    ws[f'E{current_row}'] = data2["lat"]
                    ws[f'F{current_row}'] = data2["lng"]
                elif i == len(json_data) - 1:
                    data2 = json_data[(from_chainage)]
                    ws[f'E{current_row}'] = data2["lat"]
                    ws[f'F{current_row}'] = data2["lng"]

                else:
                    data2 = json_data[(500*i + chainage_diff)]
                    ws[f'E{current_row}'] = data2["lat"]
                    ws[f'F{current_row}'] = data2["lng"]
            except KeyError:
                logger.error("End key error encountered during processing")

            if side == "LHS":
                ws[f'A{current_row}'] = from_chainage
                ws[f'B{current_row}'] = to_chainage
                ws[f'J{current_row}'] = side
            else:
                ws[f'A{current_row}'] = from_chainage
                ws[f'B{current_row}'] = to_chainage
                ws[f'J{current_row}'] = side

            # ws[f'C{current_row}'] = data["lat"]
            # ws[f'D{current_row}'] = data["lng"]
            # ws[f'E{current_row}'] = data["to_lat"]
            # ws[f'F{current_row}'] = data["to_lng"]
            try:

                # Default to 0 if "lat" doesn't exist or is None
                ws[f'C{current_row}'] = data.get("lat", 0)
                # Default to 0 if "lng" doesn't exist or is None
                ws[f'D{current_row}'] = data.get("lng", 0)
                # Default to 0 if "to_lat" doesn't exist or is None
                ws[f'E{current_row}'] = data.get("to_lat", 0)
                ws[f'F{current_row}'] = data.get("to_lng", 0)

            except KeyError as e:

                ws[f'C{current_row}'] = 0
                ws[f'D{current_row}'] = 0
                ws[f'E{current_row}'] = 0
                ws[f'F{current_row}'] = 0

            ws[f'G{current_row}'] = survey_name
            ws[f'H{current_row}'] = survey_date
            ws[f'I{current_row}'] = survey_direction

            # def get_informatory_count(section_data):
            #     return (
            #         section_data.get("INFORMATORY_SIGNS", 0) +
            #         section_data.get("DIGITAL_SPEED_DISPLAY_SIGNS", 0) +
            #         section_data.get("VARIABLE_MESSAGE_SIGNS", 0)
            #     )
            for category, col1, col2 in CATEGORY_COLUMNS:

                data1 = data.get(sides[0], {})
                data2 = data.get(sides[1], {})

                val1 = 0
                val2 = 0

                if category == "INFORMATORY_SIGNS":
                    val1 = get_informatory_count(data1)
                    val2 = get_informatory_count(data2)
                else:
                    val1 = int(data1.get(category, 0) or 0)
                    val2 = int(data2.get(category, 0) or 0)

                ws[f'{col1}{current_row}'] = val1
                ws[f'{col2}{current_row}'] = val2

        #  FOR SERVICE ROADS

        #     left_data = data.get("Left", {})
        #     right_data = data.get("Right", {})

        #     ws[f'K{current_row}'] = left_data.get("CHEVRON", 0)
        #     ws[f'L{current_row}'] = right_data.get("CHEVRON", 0)

        #     ws[f'M{current_row}'] = left_data.get(
        #         "CAUTIONARY_WARNING_SIGNS", 0)
        #     ws[f'N{current_row}'] = right_data.get(
        #         "CAUTIONARY_WARNING_SIGNS", 0)

        #     ws[f'O{current_row}'] = left_data.get("HAZARD", 0)
        #     ws[f'P{current_row}'] = right_data.get("HAZARD", 0)

        #     ws[f'Q{current_row}'] = left_data.get(
        #         "PROHIBITORY_MANDATORY_SIGNS", 0)
        #     ws[f'R{current_row}'] = right_data.get(
        #         "PROHIBITORY_MANDATORY_SIGNS", 0)

        #     ws[f'S{current_row}'] = get_informatory_count(left_data)
        #     ws[f'T{current_row}'] = get_informatory_count(right_data)

        #     # else:

        # #    FOR   MCW

        #     ws[f"K{current_row}"] = data.get("Avenue", {}).get("CHEVRON", 0)
        #     ws[f"L{current_row}"] = data.get("Median", {}).get("CHEVRON", 0)

        #     ws[f"M{current_row}"] = data.get("Avenue", {}).get(
        #         "CAUTIONARY_WARNING_SIGNS", 0)
        #     ws[f"N{current_row}"] = data.get("Median", {}).get(
        #         "CAUTIONARY_WARNING_SIGNS", 0)

        #     ws[f"O{current_row}"] = data.get("Avenue", {}).get("HAZARD", 0)
        #     ws[f"P{current_row}"] = data.get("Median", {}).get("HAZARD", 0)

        #     ws[f"Q{current_row}"] = data.get("Avenue", {}).get(
        #         "PROHIBITORY_MANDATORY_SIGNS", 0)
        #     ws[f"R{current_row}"] = data.get("Median", {}).get(
        #         "PROHIBITORY_MANDATORY_SIGNS", 0)

        #     ws[f"S{current_row}"] = get_informatory_count(
        #         data.get("Avenue", {}))
        #     ws[f"T{current_row}"] = get_informatory_count(
        #         data.get("Median", {}))

            FONT = "Microsoft Sans Serif"
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
                cell = ws[f'{col}{current_row}']
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.font = Font(name=FONT, size=10, bold=False)
                cell.border = thin_border

        # output_file_path = os.path.join(output_folder, f"{survey_data['road_name']}_formatted.xlsx")
        output_file_path = os.path.join(
            output_folder, f"{road_id}_formatted.xlsx")
        wb.save(output_file_path)

    # road_data = fetch_road_data(road_id, logger)

    if road_data:
        try:
            road_name = road_data["road"]["name"]

            start_chainage = road_data["road"]["start_chainage"]
            end_chainage = road_data["road"]["end_chainage"]

            logger.info(
                "Processing started | start_chainage=%s | end_chainage=%s",
                start_chainage,
                end_chainage)

            if start_chainage is not None:

                start_chainage = float(start_chainage)
                end_chainage = float(end_chainage)

                logger.debug("Entered inner condition block")

                chainage_diff = end_chainage - start_chainage

                logger.debug("Chainage difference: %s", chainage_diff)

                if chainage_diff < 0:

                    side = "RHS"
                    next_value = previous_value_divisible_by_500(
                        int(start_chainage))
                    value_differnce = int(start_chainage) - next_value
                    logger.debug(
                        "Next value computed | next_value=%s | value_difference=%s",
                        next_value, value_differnce)

                    if value_differnce == 0:
                        value_differnce = 500

                else:
                    side = "LHS"
                    next_value = previous_next_divisible_by_500(
                        int(start_chainage))
                    value_differnce = next_value - int(start_chainage)
                    logger.debug("Entered LHS branch (inner condition)")

                    if value_differnce == 0:
                        value_differnce = 500

            else:
                side = "LHS"
                chainage_diff = 0
                next_value = 500
                value_differnce = 500

                logger.debug("Entered outer LHS condition block")

            logger.debug(
                "Computed values | side=%s | chainage_diff=%s | next_value=%s | value_difference=%s",
                side,
                chainage_diff,
                next_value,
                value_differnce)

            mc_name = road_data["road"]["assigned_to"]["username"]
            survey_name = road_data["road"]["name"]
            survey_date = road_data["created_at"].split("T")[0]
            sub_division_name = road_data["road"]["assigned_to"]["sub_division"]
            # print(road_data["road"]["LHR_side"] , "checking ")

            # if  road_data["road"]["LHR_side"]:
            #     side = "LHS"
            # else:
            #     side = "RHS"

            logger.info("Side determined | road_id=%s | side=%s",
                        road_id, side)

        except KeyError as e:
            logger.exception("KeyError during road data parsing: %s", e)
            return

        survey_data = {
            "road_name": road_name,
            "start_chainage": start_chainage,
            "end_chainage": end_chainage,
            "end_chainage": end_chainage,
            "mc": {"name": mc_name},
            "sub_division": {"sub_division": sub_division_name},
            "survey_date": survey_date,
            "survey_name": survey_name,
            "side": side
        }

        gpx_file_path = gpx_json_path

        if not os.path.exists(gpx_file_path):
            logger.warning(
                "GPX JSON not found | road_id=%s | path=%s", road_id, gpx_file_path)
            gpx_data = {}
        else:
            with open(gpx_file_path, 'r') as f:
                gpx_data = json.load(f)

        road_json_path = output_json_path

        if not os.path.exists(road_json_path):
            logger.error(
                "Road JSON not found | road_id=%s | path=%s",
                road_id,
                road_json_path
            )
            return {
                "road_id": road_id,
                "status": "json_missing"
            }

        with open(road_json_path, 'r') as f:
            detections_data = json.load(f)

        detections_data['assets'] = sort_assets_by_distance(detections_data)

        max_distance = get_max_distance(gpx_data)

        if max_distance <= 0:
            max_distance = max(
                [float(x.get("Distance", 0) or 0)
                 for x in detections_data.get("assets", [])],
                default=0
            )

        # max_distance_ns = get_max_distance(gpx_data)

        # max_distance = previous_next_divisible_by_500(max_distance_ns)
        logger.info(
            f"Max distance before parse | road_id={road_id} | max_distance={max_distance} | diff={value_differnce}")

        # max_distance2 = survey_data.get('end_chainage', None)
        parsed_result = parse_json(
            detections_data, max_distance, gpx_data, road_name, value_differnce)

        logger.info(
            "Parse completed | road_id=%s | max_distance=%s",
            road_id,
            max_distance
        )

        # print("parsed results ..." , parsed_result )
        COUNT_JSON = os.path.join(output_folder, f"count_{road_id}.json")

        with open(COUNT_JSON, 'w') as f:
            json.dump(parsed_result, f, indent=4)

        # print("surve" , survey_data["side"])
        chainageWiseCounting(survey_data, parsed_result, side, value_differnce)
        output_excel = os.path.join(output_folder, f"{road_id}_formatted.xlsx")

        logger.info("excel3 completed successfully | road_id=%s", road_id)

        return {
            "road_id": road_id,
            "status": "success",
            "count_json": COUNT_JSON,
            "output_excel": output_excel,
        }
