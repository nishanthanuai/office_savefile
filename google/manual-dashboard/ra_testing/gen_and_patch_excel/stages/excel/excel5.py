# import os
# import json
# import requests
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
# from openpyxl.drawing.image import Image


# HANUAI_LOGO_PATH = "images/HanuAI.png"
# ROADATHENA_LOGO_PATH = "images/RA-logo-1.png"


# # --------------------------------------------------
# # API
# # --------------------------------------------------

# def fetch_road_data(roadId):
#     api_url = f"https://ndd.roadathena.com/api/surveys/roads/{roadId}"
#     response = requests.get(
#         api_url, headers={"Security-Password": "admin@123"}
#     )
#     if response.status_code == 200:
#         return response.json()
#     return None


# # --------------------------------------------------
# # COMMON HEADER FORMAT
# # --------------------------------------------------

# def apply_common_formatting(sheet, survey_data, road_data):

#     thick = Border(
#         left=Side(style="medium"),
#         right=Side(style="medium"),
#         top=Side(style="medium"),
#         bottom=Side(style="medium"),
#     )

#     sheet["A1"] = (
#         "AI Based road condition assessment detailed report by ROAD ATHENA\n"
#         f"RO : {road_data['road']['ho']['name']}\n"
#         f"PIU : {road_data['road']['ro']['name']}"
#     )

#     sheet.merge_cells("A1:I1")
#     sheet["A1"].font = Font(b=True, size=14)
#     sheet["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
#     sheet["A1"].border = thick

#     sheet["A2"] = f"Date of survey : {survey_data['survey_date']}"
#     sheet.merge_cells("A2:I2")
#     sheet["A2"].alignment = Alignment(horizontal="center")
#     sheet["A2"].border = thick

#     sheet["A3"] = f"Segment name : {survey_data['road_name']}"
#     sheet.merge_cells("A3:I3")
#     sheet["A3"].font = Font(b=True)
#     sheet["A3"].alignment = Alignment(horizontal="center")
#     sheet["A3"].border = thick

#     sheet["A4"] = "Start Chainage"
#     sheet["B4"] = survey_data["start_chainage"]
#     sheet["A5"] = "End Chainage"
#     sheet["B5"] = survey_data["end_chainage"]

#     for r in [4, 5]:
#         for c in ["A", "B"]:
#             sheet[f"{c}{r}"].alignment = Alignment(horizontal="center")
#             sheet[f"{c}{r}"].border = thick

#     sheet.add_image(Image(ROADATHENA_LOGO_PATH), "A1")
#     sheet.add_image(Image(HANUAI_LOGO_PATH), "H1")


# # --------------------------------------------------
# # MAIN EXCEL5 LOGIC
# # --------------------------------------------------

# def process_json_data5(output_json_path, output_folder, roadId):

#     json_path = os.path.join(output_json_path, f"road_{roadId}.json")
#     with open(json_path, "r") as f:
#         json_data = json.load(f)

#     road_data = fetch_road_data(roadId)
#     if not road_data:
#         print("Failed to fetch road data")
#         return

#     survey_data = {
#         "road_name": road_data["road"]["name"],
#         "start_chainage": road_data["road"]["start_chainage"],
#         "end_chainage": road_data["road"]["end_chainage"],
#         "survey_date": road_data["created_at"].split("T")[0],
#     }

#     wb_path = os.path.join(output_folder, f"{roadId}_formatted.xlsx")
#     wb = load_workbook(wb_path)

#     sheet_name = "Damaged Signs"
#     if sheet_name in wb.sheetnames:
#         del wb[sheet_name]

#     ws = wb.create_sheet(sheet_name)

#     thin = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin"),
#     )

#     headers = [
#         "Damage No",
#         "Timestamp",
#         "Asset / Anomaly type",
#         "Side",
#         "Category",
#         "Latitude",
#         "Longitude",
#         "Distance (m)",
#         "Image",
#     ]

#     for col, title in enumerate(headers, start=1):
#         cell = ws.cell(row=6, column=col, value=title)
#         cell.font = Font(b=True)
#         cell.alignment = Alignment(horizontal="center", wrap_text=True)
#         cell.fill = PatternFill("solid", fgColor="C0D3EB")
#         cell.border = thin

#     row_idx = 7
#     serial = 1

#     for item in json_data.get("assets", []) + json_data.get("anomalies", []):

#         asset_type = item.get("Asset type") or item.get("Anomaly type")

#         if asset_type != "DAMAGED_SIGN":
#             continue

#         ws[f"A{row_idx}"] = serial
#         ws[f"B{row_idx}"] = item.get("Timestamp on processed video")
#         ws[f"C{row_idx}"] = asset_type
#         ws[f"D{row_idx}"] = item.get("Side")
#         ws[f"E{row_idx}"] = item.get("category")
#         ws[f"F{row_idx}"] = item.get("Latitude")
#         ws[f"G{row_idx}"] = item.get("Longitude")
#         ws[f"H{row_idx}"] = item.get("Distance")
#         ws[f"I{row_idx}"] = item.get("image")

#         for c in range(1, 10):
#             ws.cell(row=row_idx, column=c).alignment = Alignment(
#                 horizontal="center", wrap_text=True
#             )
#             ws.cell(row=row_idx, column=c).border = thin

#         serial += 1
#         row_idx += 1

#     apply_common_formatting(ws, survey_data, road_data)

#     wb.save(wb_path)
#     print(f"Damaged Signs sheet added to {wb_path}")

# --------------------------------------------------------------------------------------

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import json
import os
import requests
# import streamlit as st
from openpyxl.drawing.image import Image
from utils.road_resolver import resolve_road_file
import logging
logger = logging.getLogger(__name__)


_IMAGES_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "images")
HANUAI_LOGO_PATH = os.path.abspath(os.path.join(_IMAGES_DIR, "HanuAI.png"))
ROADATHENA_LOGO_PATH = os.path.abspath(os.path.join(_IMAGES_DIR, "RA-logo-1.png"))


# def process_json_data5(output_json_path, output_folder, roadId):

#     road_id = roadId

# def fetch_road_data(road_id):
#     api_url = f"https://ndd.roadathena.com/api/surveys/roads/{road_id}"
#     password = os.getenv("ROAD_API_PASSWORD", "admin@123")

#     if not password:
#         logger.error("ROAD_API_PASSWORD environment variable not set")
#         return None

#     try:
#         response = requests.get(
#             api_url,
#             headers={"Security-Password": password},
#             timeout=30
#         )
#         response.raise_for_status()

#         logger.debug("Road API success | road_id=%s", road_id)
#         return response.json()

#     except requests.exceptions.RequestException:
#         logger.exception("API request failed | road_id=%s", road_id)
#         return None


def apply_common_formatting(sheet, survey_data, road_data):

    thick_border = Border(left=Side(style='medium'),
                          right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # A1 - Title
    # sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nMC: {survey_data['mc']['name']}\nSubdivision: {survey_data['sub_division']['sub_division']}"
    sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {road_data['road']['ho']['name']}\nPIU : {road_data['road']['ro']['name']}"
    sheet['A1'].font = Font(b=True, size=14)
    sheet['A1'].alignment = Alignment(wrap_text=True, horizontal='center')
    sheet.merge_cells('A1:M1')
    sheet['A1'].border = thick_border

    for col in range(1, 14):  # Columns A to L
        cell = sheet.cell(row=1, column=col)
        cell.border = thick_border

    # A2 - Date of survey
    sheet["A2"] = f"Date of survey: {survey_data['survey_date']}"
    sheet['A2'].font = Font(size=13)
    sheet.merge_cells('A2:K2')
    sheet['A2'].alignment = Alignment(horizontal='center')
    sheet['A2'].border = thick_border

    for col in range(1, 12):  # Columns A to L
        cell = sheet.cell(row=2, column=col)
        cell.border = thick_border

    # A3 - Segment name
    sheet["A3"] = f"Segment name: {survey_data['road_name']}"
    sheet.merge_cells('A3:K3')  # Merge cells A3 to E3
    sheet['A3'].font = Font(size=13, b=True)
    sheet['A3'].alignment = Alignment(horizontal='center')
    sheet['A3'].border = thick_border

    for col in range(1, 12):  # Columns A to L
        cell = sheet.cell(row=3, column=col)
        cell.border = thick_border

    # A4 - Start Chainage
    sheet["A4"] = f"Start Chainage"
    sheet['A4'].font = Font(size=13, b=True)
    sheet['A4'].alignment = Alignment(horizontal='center')
    sheet['A4'].border = thick_border

    # B4 - Start Chainage value
    sheet["B4"] = f"{survey_data['start_chainage']}"
    sheet['B4'].font = Font(size=13)
    sheet['B4'].alignment = Alignment(horizontal='center')
    sheet['B4'].border = thick_border

    # A5 - End Chainage
    sheet["A5"] = f"End Chainage"
    sheet['A5'].font = Font(size=13, b=True)
    sheet['A5'].alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')
    sheet['A5'].border = thick_border

    # B5 - End Chainage value
    sheet["B5"] = f"{survey_data['end_chainage']}"
    sheet['B5'].font = Font(size=13)
    sheet['B5'].alignment = Alignment(
        wrap_text=True, horizontal='center', vertical='center')
    sheet['B5'].border = thick_border

    img = Image(ROADATHENA_LOGO_PATH)
    img.anchor = 'A1'
    sheet.add_image(img)

    img = Image(HANUAI_LOGO_PATH)
    img.anchor = 'K1'
    sheet.add_image(img)

    sheet.row_dimensions[1].height = 60
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 30
    sheet.row_dimensions[4].height = 50
    sheet.row_dimensions[5].height = 50
    sheet.row_dimensions[6].height = 50

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20

    return sheet


def create_detailed_report(data, survey_data, output_folder, road_id, road_data):
    wb = load_workbook(os.path.join(
        output_folder, f"{road_id}_formatted.xlsx"))

    # Check and create "Encroachment Signs" sheet
    encroachment_sheet_name = "Damaged Signs"
    if encroachment_sheet_name in wb.sheetnames:
        del wb[encroachment_sheet_name]
    ws_encroachment = wb.create_sheet(encroachment_sheet_name)
    ws_encroachment.title = encroachment_sheet_name

    # Formatting for Encroachment Signs sheet
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws_encroachment["A6"] = "Damaged Sign Number"
    ws_encroachment["B6"] = "Timestamp on processed video"
    ws_encroachment["C6"] = "Anomaly type"
    ws_encroachment["D6"] = "Side"
    ws_encroachment["E6"] = "category"
    ws_encroachment["F6"] = "Latitude"
    ws_encroachment["G6"] = "Longitude"
    ws_encroachment["H6"] = "Distance from start point in meters"
    ws_encroachment["I6"] = "Image Link"

    for col in range(1, 10):
        cell = ws_encroachment.cell(row=6, column=col)
        cell.font = Font(b=True, size=12)
        cell.alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(
            start_color='c0d3eb', end_color='c0d3eb', patternType='solid')

    # Filtered data
    filtered_data = [
        item for item in (data.get("assets", []) + data.get("anomalies", []))
        if (
            item.get("Asset type") in ["DAMAGED_SIGN"]
            or item.get("Anomaly type") in ["DAMAGED_SIGN"]
        )
    ]

    # Populate the "Encroachment Signs" sheet
    serial_number = 1
    for i, row_item in enumerate(filtered_data, start=7):
        ws_encroachment[f'A{i}'] = serial_number
        ws_encroachment[f'B{i}'] = row_item["Timestamp on processed video"]
        ws_encroachment[f'C{i}'] = row_item.get(
            "Asset type") or row_item.get("Anomaly type")
        ws_encroachment[f'D{i}'] = row_item.get(
            "Side") or row_item.get("side")
        ws_encroachment[f'E{i}'] = row_item.get(
            "Category") or row_item.get("category")
        ws_encroachment[f'F{i}'] = row_item["Latitude"]
        ws_encroachment[f'G{i}'] = row_item["Longitude"]
        ws_encroachment[f'H{i}'] = row_item["Distance"]
        ws_encroachment[f'I{i}'] = row_item["image"]

        for col in range(1, 10):
            cell = ws_encroachment.cell(row=i, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                wrap_text=True, horizontal='center', vertical='center')

        serial_number += 1

    apply_common_formatting(ws_encroachment, survey_data, road_data)

    # Save workbook
    output_file_path = os.path.join(
        output_folder, f"{road_id}_formatted.xlsx")
    wb.save(output_file_path)
    logger.info(
        "Damaged Signs sheet added successfully | road_id=%s | path=%s",
        road_id,
        output_file_path
    )


# Load JSON data
# Resolve correct road JSON (mcw / service)
def run(output_json_path: str, output_folder: str, road_id: int, road_data:Dict ):
    """
    Master entry point for excel5.
    Creates Damaged Signs sheet in {road_id}_formatted.xlsx
    """

    logger.info("Starting excel5 | road_id=%s", road_id)

    # Resolve correct road JSON file (mcw/service)
    road_json_path = output_json_path
    if not road_json_path:
        logger.error(
            "road_%s.json not found in mcw/service folders",
            road_id
        )
        return {"road_id": road_id, "status": "json_missing"}

    # Load JSON data
    try:
        with open(road_json_path, "r") as file:
            json_data = json.load(file)
    except Exception as e:
        logger.exception(
            "Failed to load road JSON | road_id=%s | error=%s",
            road_id,
            e
        )
        return {"road_id": road_id, "status": "json_load_failed"}

    # Fetch road metadata from API
    # road_data = fetch_road_data(road_id)

    if not road_data:
        logger.error(
            "Failed to fetch road data. Exiting process | road_id=%s",
            road_id
        )
        return {"road_id": road_id, "status": "api_failed"}

    # Extract metadata safely
    try:
        survey_data = {
            "road_name": road_data["road"]["name"],
            "start_chainage": road_data["road"]["start_chainage"],
            "end_chainage": road_data["road"]["end_chainage"],
            "survey_date": road_data["created_at"].split("T")[0],
        }
    except KeyError as e:
        logger.exception(
            "KeyError while extracting road metadata | road_id=%s | error=%s",
            road_id,
            e
        )
        return {"road_id": road_id, "status": "metadata_error"}

    # Ensure formatted workbook exists
    formatted_path = os.path.join(output_folder, f"{road_id}_formatted.xlsx")

    if not os.path.exists(formatted_path):
        logger.error(
            "Formatted workbook not found before excel5 | road_id=%s | path=%s",
            road_id,
            formatted_path
        )
        return {"road_id": road_id, "status": "workbook_missing"}

    # Create Damaged Signs sheet
    try:
        create_detailed_report(
            json_data,
            survey_data,
            output_folder,
            road_id, road_data
        )
    except Exception as e:
        logger.exception(
            "Failed while creating Damaged Signs sheet | road_id=%s | error=%s",
            road_id,
            e
        )
        return {"road_id": road_id, "status": "sheet_creation_failed"}

    logger.info("excel5 completed successfully | road_id=%s", road_id)

    return {
        "road_id": road_id,
        "status": "success",
        "output_excel": formatted_path
    }
