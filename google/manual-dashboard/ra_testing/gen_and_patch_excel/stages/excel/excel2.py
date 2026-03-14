# =================================================================================================================================
# =================================================================================================================================
import os
import json
import logging
import requests
from typing import Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.drawing.image import Image
# from utils.road_resolver import resolve_road_file

logger = logging.getLogger(__name__)


_IMAGES_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "images")
HANUAI_LOGO_PATH = os.path.abspath(os.path.join(_IMAGES_DIR, "HanuAI.png"))
ROADATHENA_LOGO_PATH = os.path.abspath(os.path.join(_IMAGES_DIR, "RA-logo-1.png"))


def run(output_json_path: str, output_folder: str, road_id: int,road_data:Dict) -> Dict:

    # def fetch_road_data(road_id):
    #     api_url = f"https://ndd.roadathena.com/api/surveys/roads/{road_id}"
    #     # Use environment variable for password
    #     # fallback to default if not set
    #     password = os.getenv("ROAD_API_PASSWORD", "admin@123")
    #     response = requests.get(api_url, headers={"Security-Password": password})

    #     if response.status_code == 200:
    #         road_data = response.json()
    #         return road_data
    #     else:
    #         logger.error(
    #             f"API fetch failed | road_id={road_id} | status={response.status_code}"
    #         )
    #         return None

    def apply_common_formatting(asset_sheet, survey_data, road_data):

        thick_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )

        # A1 - Title
        # sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nMC: {survey_data['mc']['name']}\nSubdivision: {survey_data['sub_division']['sub_division']}"
        asset_sheet["A1"] = (
            f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {road_data['road']['ho']['name']}\nPIU : {road_data['road']['ro']['name']}"
        )
        asset_sheet["A1"].font = Font(b=True, size=14)
        asset_sheet["A1"].alignment = Alignment(wrap_text=True, horizontal="center")
        asset_sheet.merge_cells("A1:J1")
        asset_sheet["A1"].border = thick_border
        for col in range(1, 11):  # Columns A to I
            cell = asset_sheet.cell(row=1, column=col)
            cell.border = thick_border

        # A2 - Date of survey
        asset_sheet["A2"] = f"Date of survey: {survey_data['survey_date']}"
        asset_sheet["A2"].font = Font(size=13)
        asset_sheet.merge_cells("A2:I2")
        asset_sheet["A2"].alignment = Alignment(horizontal="center")
        asset_sheet["A2"].border = thick_border

        for col in range(1, 10):  # Columns A to I
            cell = asset_sheet.cell(row=2, column=col)
            cell.border = thick_border

        # A3 - Segment name
        asset_sheet["A3"] = f"Segment name: {survey_data['road_name']}"
        asset_sheet.merge_cells("A3:I3")  # Merge cells A3 to E3
        asset_sheet["A3"].font = Font(size=13, b=True)
        asset_sheet["A3"].alignment = Alignment(horizontal="center")
        asset_sheet["A3"].border = thick_border

        for col in range(1, 10):  # Columns A to I
            cell = asset_sheet.cell(row=3, column=col)
            cell.border = thick_border

        # A4 - Start Chainage
        asset_sheet["A4"] = f"Start Chainage"
        asset_sheet["A4"].font = Font(size=10, b=True)
        asset_sheet["A4"].alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        asset_sheet["A4"].border = thick_border

        # B4 - Start Chainage value
        asset_sheet["B4"] = f"{survey_data['start_chainage']}"
        asset_sheet["B4"].font = Font(size=15)
        asset_sheet["B4"].alignment = Alignment(horizontal="center")
        asset_sheet["B4"].border = thick_border

        # A5 - End Chainage
        asset_sheet["A5"] = f"End Chainage"
        asset_sheet["A5"].font = Font(size=10, b=True)
        asset_sheet["A5"].alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        asset_sheet["A5"].border = thick_border

        # B5 - End Chainage value
        asset_sheet["B5"] = f"{survey_data['end_chainage']}"
        asset_sheet["B5"].font = Font(size=15)
        asset_sheet["B5"].alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        asset_sheet["B5"].border = thick_border

        img = Image(ROADATHENA_LOGO_PATH)
        img.anchor = "A1"
        asset_sheet.add_image(img)
        img = Image(HANUAI_LOGO_PATH)
        img.anchor = "I1"
        asset_sheet.add_image(img)

        asset_sheet.row_dimensions[1].height = 60
        asset_sheet.row_dimensions[2].height = 30
        asset_sheet.row_dimensions[3].height = 40
        asset_sheet.row_dimensions[4].height = 50
        asset_sheet.row_dimensions[5].height = 50
        asset_sheet.row_dimensions[6].height = 50
        asset_sheet.column_dimensions["A"].width = 20
        asset_sheet.column_dimensions["B"].width = 20

        return asset_sheet

    def create_detailed_report(data, survey_data, road_data, output_folder, road_id):
        # Create a fresh workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # --- 1. ASSETS SHEET ---
        logger.info(f"Creating Assets sheet | road_id={road_id}")

        asset_sheet = workbook.create_sheet(title="Assets")
        apply_common_formatting(asset_sheet, survey_data, road_data)

        # Asset Headers (Row 6)
        asset_headers = [
            "Sr No.",
            "Timestamp on processed video",
            "Asset type",
            "Side",
            "Category",
            "Latitude",
            "Longitude",
            "Distance from start point (m)",
            "Image Link",
        ]

        for col_num, header in enumerate(asset_headers, 1):
            cell = asset_sheet.cell(row=6, column=col_num, value=header)
            cell.font = Font(b=True, size=11)
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            cell.border = thin_border
            cell.fill = PatternFill(
                start_color="c0d3eb", end_color="c0d3eb", patternType="solid"
            )

        # --- MODIFICATION START ---
        # 1. Define the types to exclude
        excluded_asset_types = [
            "ADVERTISEMENT_ENCHROACHMENT_SIGNS",
            "NON_STANDARD_INFOMATORY_SIGNS",
        ]

        # 2. Filter the assets list to exclude those types
        assets = data.get("assets", [])
        filtered_assets = [
            item
            for item in assets
            if item.get("Asset type") not in excluded_asset_types
        ]
        # --- MODIFICATION END ---

        # Asset Data (Starting Row 7) - Using the filtered list
        for i, item in enumerate(filtered_assets, 1):
            row_idx = i + 6
            asset_sheet.cell(row=row_idx, column=1, value=i)
            asset_sheet.cell(
                row=row_idx, column=2, value=item.get("Timestamp on processed video")
            )
            asset_sheet.cell(row=row_idx, column=3, value=item.get("Asset type"))
            asset_sheet.cell(row=row_idx, column=4, value=item.get("Side"))
            category_val = (
                item.get("category")
                if item.get("category") is not None
                else item.get("Category")
            )
            asset_sheet.cell(row=row_idx, column=5, value=category_val)
            asset_sheet.cell(row=row_idx, column=6, value=item.get("Latitude"))
            asset_sheet.cell(row=row_idx, column=7, value=item.get("Longitude"))
            asset_sheet.cell(row=row_idx, column=8, value=item.get("Distance"))

            # Hyperlink for Assets (Column I)
            img_url = item.get("image", "")
            cell_url = asset_sheet.cell(row=row_idx, column=9, value=img_url)
            if img_url:
                cell_url.hyperlink = img_url
                cell_url.font = Font(color="0000FF", underline="single")

            # Apply formatting to all cells in the row (Columns 1 to 9)
            for col in range(1, 10):
                cell = asset_sheet.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )

        # Set specific width for the Image Link column
        asset_sheet.column_dimensions["I"].width = 60

        # --- 2. ANOMALIES SHEET ---
        logger.info(f"Creating Anomalies sheet | road_id={road_id}")

        anomaly_sheet = workbook.create_sheet(title="Anomalies")
        apply_common_formatting(anomaly_sheet, survey_data, road_data)

        # Anomaly Headers (Row 6)
        anomaly_headers = [
            "Sr No.",
            "Timestamp on processed video",
            "Anomaly Type",
            "Latitude",
            "Longitude",
            "Distance (m)",
            "Image Link",
        ]

        for col_num, header in enumerate(anomaly_headers, 1):
            cell = anomaly_sheet.cell(row=6, column=col_num, value=header)
            cell.font = Font(b=True, size=11)
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            cell.border = thin_border
            cell.fill = PatternFill(
                start_color="c0d3eb", end_color="c0d3eb", patternType="solid"
            )

        # Anomaly Data (Starting Row 7)
        anomalies = data.get("anomalies", [])
        for i, item in enumerate(anomalies, 1):
            row_idx = i + 6
            anomaly_sheet.cell(row=row_idx, column=1, value=i)
            anomaly_sheet.cell(
                row=row_idx, column=2, value=item.get("Timestamp on processed video")
            )
            anomaly_sheet.cell(row=row_idx, column=3, value=item.get("Anomaly type"))
            anomaly_sheet.cell(row=row_idx, column=4, value=item.get("Latitude"))
            anomaly_sheet.cell(row=row_idx, column=5, value=item.get("Longitude"))
            anomaly_sheet.cell(row=row_idx, column=6, value=item.get("Distance"))

            img_url = item.get("image", "")
            cell_url = anomaly_sheet.cell(row=row_idx, column=7, value=img_url)
            if img_url:
                cell_url.hyperlink = img_url
                cell_url.font = Font(color="0000FF", underline="single")

            # Apply formatting to all cells in the row
            for col in range(1, 8):
                cell = anomaly_sheet.cell(row=row_idx, column=col)
                cell.border = thin_border
                # wrap_text=True is the key here
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )

        # --- Column Width Adjustments ---
        for sheet in [asset_sheet, anomaly_sheet]:
            sheet.column_dimensions["A"].width = 10
            sheet.column_dimensions["B"].width = 25
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 15
            sheet.column_dimensions["E"].width = 15
            sheet.column_dimensions["F"].width = 15
            sheet.column_dimensions["G"].width = 15
            if sheet.title == "Assets":
                sheet.column_dimensions["H"].width = 20
                sheet.column_dimensions["I"].width = 15

        # Final Save
        os.makedirs(output_folder, exist_ok=True)
        output_file_path = os.path.join(output_folder, f"{road_id}_formatted.xlsx")
        workbook.save(output_file_path)
        logger.info(f"Excel generated | road_id={road_id} | file={output_file_path}")
        return output_file_path

    # Main logic
    try:
        # road_data = fetch_road_data(road_id)
        file_path = output_json_path

        if not os.path.exists(file_path):
            logger.error(f"JSON not found | road_id={road_id} | path={file_path}")
            return {"road_id": road_id, "status": "json_missing"}
        with open(file_path, "r", encoding="utf-8") as file:
            json_data = json.load(file)

        if not road_data:
            logger.error(f"Missing API data | road_id={road_id}")
            return {"road_id": road_id, "status": "api_missing"}

        survey_data = {
            "road_name": road_data["road"].get("name", "N/A"),
            "start_chainage": road_data["road"].get("start_chainage", 0),
            "end_chainage": road_data["road"].get("end_chainage", 0),
            "survey_date": road_data.get("created_at", "").split("T")[0],
        }

        output_file = create_detailed_report(
            json_data, survey_data, road_data, output_folder, road_id
        )

        return {"road_id": road_id, "status": "success", "output_file": output_file}

    except Exception as e:
        logger.exception(f"Excel3 failed | road_id={road_id}")
        raise
